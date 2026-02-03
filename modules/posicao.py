import streamlit as st
import pandas as pd
import numpy as np
import re
import io
import pydeck as pdk
import reverse_geocoder as rg
import pycountry
from modules.processar_relatorio import extrair_odometros
from modules.utils import convert_df_to_csv

def to_excel(df, highlight_col=None, highlight_value='Sim'):
    """Converts a dataframe to an Excel file in-memory, with optional highlighting."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Manutenção')
        
        if highlight_col and highlight_col in df.columns:
            from openpyxl.styles import PatternFill
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Manutenção']
            
            # Define fill style for highlighting
            highlight_fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid') # Light Yellow

            # Iterate through rows and apply formatting if the condition is met
            for row_idx, value in enumerate(df[highlight_col], start=2): # start=2 for 1-based index + header
                if value == highlight_value:
                    for col_idx in range(1, len(df.columns) + 1):
                        worksheet.cell(row=row_idx, column=col_idx).fill = highlight_fill
                        
    return output.getvalue()

def standardize_column_names(df):
    """Padroniza os nomes das colunas para garantir a compatibilidade."""
    column_map = {
        'modelo de hw': 'Modelo de HW', 'modelo do hw': 'Modelo de HW', 'modelo hw': 'Modelo de HW',
        'modelo hardware': 'Modelo de HW', 'hardware': 'Modelo de HW', 'modelo': 'Modelo de HW',
        'serial': 'Serial', 'serial number': 'Serial', 'device id': 'Serial', 'numero de serie': 'Serial',
    }
    rename_dict = {col: column_map[col.lower()] for col in df.columns if isinstance(col, str) and col.lower() in column_map}
    if rename_dict:
        df.rename(columns=rename_dict, inplace=True)
    return df

def _normalize_key(value):
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return None
    text = str(value).strip()
    return text.upper() if text else None

def _prepare_posicao_odometro_df(df):
    if df is None or df.empty:
        return pd.DataFrame()
    cols = ['Serial', 'Placa', 'odometro', 'odometro_can']
    existing_cols = [c for c in cols if c in df.columns]
    df_copy = df[existing_cols].copy()

    df_copy['serial_key'] = df_copy.get('Serial').apply(_normalize_key) if 'Serial' in df_copy else None
    df_copy['placa_key'] = df_copy.get('Placa').apply(_normalize_key) if 'Placa' in df_copy else None
    df_copy['__pos_index__'] = df_copy.index
    return df_copy

def _prepare_cps_odometro_df(df):
    if df is None or df.empty or 'Evento / Ignição' not in df.columns:
        return None
    cps_copy = df.copy()
    cps_copy['Evento / Ignição'] = cps_copy['Evento / Ignição'].astype(str)

    odometer_values = cps_copy['Evento / Ignição'].apply(
        lambda txt: pd.Series(extrair_odometros(txt), index=['odometro', 'odometro_can'])
    )
    cps_copy = pd.concat([cps_copy, odometer_values], axis=1)

    splitted = cps_copy['Evento / Ignição'].str.split('/', n=1, expand=True)
    cps_copy['Evento'] = splitted[0].str.strip()
    if splitted.shape[1] > 1:
        cps_copy['Ignição'] = splitted[1].str.strip()
    else:
        cps_copy['Ignição'] = None

    serial_col = next((c for c in cps_copy.columns if 'serial' in c.lower()), None)
    placa_col = next((c for c in cps_copy.columns if 'placa' in c.lower()), None)

    cps_copy['Serial'] = cps_copy[serial_col] if serial_col else None
    cps_copy['Placa'] = cps_copy[placa_col] if placa_col else None

    cps_copy['serial_key'] = cps_copy['Serial'].apply(_normalize_key)
    cps_copy['placa_key'] = cps_copy['Placa'].apply(_normalize_key)
    cps_copy['__cps_index__'] = cps_copy.index

    return cps_copy

def _merge_on_key(df_pos, df_cps, key, match_label):
    if key not in df_pos.columns or key not in df_cps.columns:
        return pd.DataFrame()
    df_pos_key = df_pos[df_pos[key].notna()].copy()
    df_cps_key = df_cps[df_cps[key].notna()].copy()
    if df_pos_key.empty or df_cps_key.empty:
        return pd.DataFrame()
    merged = pd.merge(
        df_pos_key,
        df_cps_key,
        on=key,
        suffixes=('_pos', '_cps'),
        how='inner',
        sort=False,
    )
    if merged.empty:
        return merged
    merged['Match via'] = match_label
    return merged

def cruzar_odometros_posicao_cps(df_pos, df_cps):
    df_cps_prepared = _prepare_cps_odometro_df(df_cps)
    if df_cps_prepared is None or df_pos is None or df_pos.empty:
        return None

    df_pos_prepared = _prepare_posicao_odometro_df(df_pos)
    if df_pos_prepared.empty:
        return pd.DataFrame()

    serial_merge = _merge_on_key(df_pos_prepared, df_cps_prepared, 'serial_key', 'Serial')

    matched_pos = set(serial_merge['__pos_index__']) if not serial_merge.empty else set()
    matched_cps = set(serial_merge['__cps_index__']) if not serial_merge.empty else set()

    pos_remaining = df_pos_prepared[~df_pos_prepared['__pos_index__'].isin(matched_pos)]
    cps_remaining = df_cps_prepared[~df_cps_prepared['__cps_index__'].isin(matched_cps)]

    placa_merge = _merge_on_key(pos_remaining, cps_remaining, 'placa_key', 'Placa')

    final = pd.concat([serial_merge, placa_merge], ignore_index=True, sort=False)
    if final.empty:
        return final

    rename_map = {
        'Serial_pos': 'Serial (Posição)',
        'Placa_pos': 'Placa (Posição)',
        'Serial_cps': 'Serial (CPS)',
        'Placa_cps': 'Placa (CPS)',
        'odometro_pos': 'Odômetro Posição',
        'odometro_can_pos': 'Odômetro CAN Posição',
        'odometro_cps': 'Odômetro CPS',
        'odometro_can_cps': 'Odômetro CAN CPS',
        'Evento / Ignição': 'Evento / Ignição (CPS)',
        'Ignição': 'Ignição (CPS)'
    }
    final = final.rename(columns=rename_map)

    cols_order = [
        'Serial (Posição)', 'Placa (Posição)', 'Odômetro Posição', 'Odômetro CAN Posição',
        'Serial (CPS)', 'Placa (CPS)', 'Odômetro CPS', 'Odômetro CAN CPS',
        'Evento / Ignição (CPS)', 'Ignição (CPS)', 'Match via'
    ]
    cols_existentes = [col for col in cols_order if col in final.columns]
    final = final[cols_existentes + [col for col in final.columns if col not in cols_existentes]]
    final.drop(columns=[c for c in ['serial_key', 'placa_key', '__pos_index__', '__cps_index__'] if c in final.columns], inplace=True, errors='ignore')
    final.fillna('N/A', inplace=True)
    return final
def map_to_canonical_model_names(df):
    """
    Agrupa nomes de modelos de HW semelhantes e os padroniza para a forma mais comum,
    preservando a formatação original (maiúsculas, hífens, etc.).
    """
    if 'Modelo de HW' not in df.columns:
        return df

    # Garante que a coluna seja do tipo string para manipulação
    df['Modelo de HW'] = df['Modelo de HW'].astype(str).replace('nan', np.nan)
    
    original_models = df['Modelo de HW'].dropna().unique()
    
    # Gera uma chave "limpa" para cada modelo (minúsculas, sem espaços/hífens)
    model_groups = {}
    for model in original_models:
        clean_key = re.sub(r'[\s_-]', '', model.lower())
        if clean_key not in model_groups:
            model_groups[clean_key] = []
        model_groups[clean_key].append(model)

    # Para cada grupo, encontra o nome canônico (o mais frequente) e cria o mapeamento
    canonical_mapping = {}
    all_series = df['Modelo de HW'].dropna()
    for _, messy_models in model_groups.items():
        if not messy_models:
            continue
        
        # Encontra a ocorrência mais comum dentro do grupo de nomes semelhantes
        most_frequent_name = all_series[all_series.isin(messy_models)].mode()
        
        # Define o nome canônico
        canonical_name = messy_models[0] # Fallback para o primeiro item
        if not most_frequent_name.empty:
            canonical_name = most_frequent_name.iloc[0]
            
        # Mapeia todas as variações para o nome canônico
        for model in messy_models:
            canonical_mapping[model] = canonical_name
            
    # Aplica o mapeamento na coluna do DataFrame
    df['Modelo de HW'] = df['Modelo de HW'].map(canonical_mapping)
    return df

def fill_missing_models(df):
    """Preenche os valores ausentes da coluna 'Modelo de HW' com base no prefixo do 'Serial'."""
    if 'Serial' not in df.columns or 'Modelo de HW' not in df.columns:
        return df, 0

    df['Serial'] = df['Serial'].astype(str)
    df['Modelo de HW'] = df['Modelo de HW'].apply(lambda x: x if isinstance(x, str) and x.strip() else np.nan)
    initial_filled_count = df['Modelo de HW'].notnull().sum()

    # Mapa dinâmico com prefixo de 5 (mais específico)
    df_com_modelo = df.dropna(subset=['Modelo de HW']).copy()
    if not df_com_modelo.empty:
        df_com_modelo['prefixo_5'] = df_com_modelo['Serial'].str[:5]
        try:
            dynamic_map_5 = df_com_modelo.groupby('prefixo_5')['Modelo de HW'].agg(lambda x: x.mode().iloc[0]).to_dict()
        except IndexError:
            dynamic_map_5 = {}
    else:
        dynamic_map_5 = {}

    # Mapa hardcoded com prefixo de 3
    hardcoded_map_3 = {'130': 'MXT-130', '202': 'MT-2000'}

    # Aplicar os mapas
    missing_mask = df['Modelo de HW'].isnull()
    if missing_mask.any():
        # 1. Tenta preencher com o mapa dinâmico de 5 prefixos
        prefixes_5 = df.loc[missing_mask, 'Serial'].str[:5]
        df.loc[missing_mask, 'Modelo de HW'] = prefixes_5.map(dynamic_map_5)

        # 2. Para os que ainda faltam, tenta preencher com o mapa hardcoded de 3 prefixos
        still_missing_mask = df['Modelo de HW'].isnull()
        if still_missing_mask.any():
            prefixes_3 = df.loc[still_missing_mask, 'Serial'].str[:3]
            # fillna para não sobrescrever o que já foi achado
            df.loc[still_missing_mask, 'Modelo de HW'] = df.loc[still_missing_mask, 'Modelo de HW'].fillna(prefixes_3.map(hardcoded_map_3))

    filled_count = df['Modelo de HW'].notnull().sum() - initial_filled_count
    return df, filled_count

def process_location_data(df):
    """
    Prioriza a fonte da posição, preenche dados de localização ausentes
    usando geocodificação reversa e normaliza as colunas de localização.
    """
    st.info("Processando e enriquecendo dados de localização...")

    # --- 1. Lógica de Prioridade da Fonte de Posição ---
    if 'tecnologia_posicao' in df.columns and 'Modelo de HW' in df.columns:
        excluded_models = ['Contingencia', 'Iotracking', 'A40a', 'A40b']
        
        conditions = [
            df['Modelo de HW'].str.contains('|'.join(excluded_models), case=False, na=False),
            df['tecnologia_posicao'] == 'GSM',
            df['tecnologia_posicao'] == 'Lorawan',
            df['tecnologia_posicao'] == 'P2P',
            df['tecnologia_posicao'] == 'P2P/LoraWan'
        ]
        choices = ['Original', 'GSM', 'LoraWan', 'P2P', 'P2P/LoraWan']
        
        df['fonte_posicao'] = np.select(conditions, choices, default='N/A')
        st.success("Fonte da posição priorizada.")
    else:
        df['fonte_posicao'] = 'N/A'

    # --- 2. Geocodificação Reversa para Dados Ausentes ---
    df['cidade_final'] = df.get('cidade')
    df['estado_final'] = df.get('estado')
    df['pais_final'] = df.get('país') # Assumindo que a coluna pode se chamar 'país'

    # Identificar linhas que precisam de geocodificação
    needs_geocoding_mask = (df['cidade_final'].isnull() | df['estado_final'].isnull()) & df['lat'].notna() & df['long'].notna()
    rows_to_geocode = df[needs_geocoding_mask]

    if not rows_to_geocode.empty:
        st.info(f"Buscando cidade/estado para {len(rows_to_geocode)} registros sem essa informação...")
        try:
            coordinates = list(zip(pd.to_numeric(rows_to_geocode['lat'], errors='coerce'), 
                                   pd.to_numeric(rows_to_geocode['long'], errors='coerce')))
            
            # Filtra coordenadas inválidas que podem ter passado
            valid_coords_with_indices = [(coord, index) for coord, index in zip(coordinates, rows_to_geocode.index) if -90 <= coord[0] <= 90 and -180 <= coord[1] <= 180]
            
            if valid_coords_with_indices:
                valid_coords = [item[0] for item in valid_coords_with_indices]
                valid_indices = [item[1] for item in valid_coords_with_indices]

                results = rg.search(valid_coords)
                
                cities = [res.get('name', '') for res in results]
                states = [res.get('admin1', '') for res in results]
                countries_iso = [res.get('cc', '') for res in results]
                
                countries = []
                for iso in countries_iso:
                    try:
                        country = pycountry.countries.get(alpha_2=iso)
                        countries.append(country.name if country else iso)
                    except:
                        countries.append(iso)

                geocoded_data = pd.DataFrame({
                    'cidade_final': cities,
                    'estado_final': states,
                    'pais_final': countries
                }, index=valid_indices)

                # Preenche os valores nulos com os dados geocodificados
                df['cidade_final'].fillna(geocoded_data['cidade_final'], inplace=True)
                df['estado_final'].fillna(geocoded_data['estado_final'], inplace=True)
                df['pais_final'].fillna(geocoded_data['pais_final'], inplace=True)
                st.success("Busca de localização finalizada.")
        
        except Exception as e:
            st.warning(f"Ocorreu um erro durante a geocodificação reversa: {e}")

    return df

def analisar_ultima_posicao(df_processado):
    """Recebe o DataFrame, aplica filtros e exibe os dados e mapas."""
    if df_processado is None or df_processado.empty:
        st.warning("Nenhum dado de posição para analisar.")
        return

    # Padroniza nomes de colunas, modelos e preenche dados ausentes
    df_processado = standardize_column_names(df_processado)
    if 'Modelo de HW' in df_processado.columns and 'Serial' in df_processado.columns:
        df_processado = map_to_canonical_model_names(df_processado)
        df_processado, filled_count = fill_missing_models(df_processado)
        if filled_count > 0:
            st.success(f"✅ {filled_count} modelos de hardware foram preenchidos automaticamente com base no serial!")
    
    # Processa e enriquece dados de localização (prioridade, geocoding)
    df_processado = process_location_data(df_processado)

    with st.expander("🔗 Cruzar com Base de Ativos (Opcional)"):
        arquivo_ativos = st.file_uploader(
            "Selecione o arquivo da base de ativos (Excel ou CSV). A base deve conter a coluna 'Serial'.",
            type=['xlsx', 'csv']
        )
        if arquivo_ativos:
            try:
                if arquivo_ativos.name.endswith('.xlsx'):
                    df_ativos = pd.read_excel(arquivo_ativos, engine='openpyxl')
                else:
                    try:
                        df_ativos = pd.read_csv(arquivo_ativos, sep=';', encoding='utf-8', on_bad_lines='warn')
                    except Exception:
                         df_ativos = pd.read_csv(arquivo_ativos, sep=',', encoding='utf-8', on_bad_lines='warn')

                st.info("Colunas detectadas na base de ativos: " + ", ".join(df_ativos.columns))
                
                df_ativos = standardize_column_names(df_ativos)

                if 'Serial' in df_ativos.columns:
                    df_processado['Serial'] = df_processado['Serial'].astype(str).str.strip()
                    df_ativos['Serial'] = df_ativos['Serial'].astype(str).str.strip()
                    df_ativos.drop_duplicates(subset=['Serial'], keep='last', inplace=True)
                    
                    processado_cols = df_processado.columns.tolist()
                    ativos_cols_to_merge = [col for col in df_ativos.columns if col not in processado_cols or col == 'Serial']
                    
                    df_processado = pd.merge(df_processado, df_ativos[ativos_cols_to_merge], on='Serial', how='left')
                    st.success("✅ Base de ativos cruzada com sucesso!")

                    # Lógica de Comparação de Região
                    df_processado['mesma_regiao'] = 'Não'
                    if all(col in df_processado.columns for col in ['cidade_final', 'estado_final', 'Cidade']) and ('Uf Proprietario' in df_processado.columns or 'Estado' in df_processado.columns):
                        
                        posicao_cidade = df_processado['cidade_final'].str.lower().str.strip()
                        posicao_estado = df_processado['estado_final'].str.lower().str.strip()
                        ativos_cidade = df_processado['Cidade'].str.lower().str.strip()
                        
                        ativos_estado_col = 'Uf Proprietario' if 'Uf Proprietario' in df_processado.columns else 'Estado'
                        ativos_estado = df_processado[ativos_estado_col].str.lower().str.strip()

                        match_condition = (posicao_cidade == ativos_cidade) & (posicao_estado == ativos_estado) & (posicao_cidade.notna())
                        df_processado.loc[match_condition, 'mesma_regiao'] = 'Sim'
                        st.info("Comparação de região concluída.")

                    if st.toggle("Mostrar apenas equipamentos na mesma região da base de ativos"):
                        if 'mesma_regiao' in df_processado.columns:
                            df_processado = df_processado[df_processado['mesma_regiao'] == 'Sim'].copy()
                            st.info(f"Exibindo {len(df_processado)} equipamentos que estão na mesma região da base.")
                        else:
                            st.warning("A comparação de região não pôde ser executada. Verifique as colunas das bases.")

                    # Filtro por cliente
                    if 'Cliente' in df_processado.columns:
                        cliente_search = st.text_input("Pesquisar por nome do cliente para filtrar a visão:", placeholder="Digite o nome do cliente")
                        if cliente_search:
                            df_filtrado_cliente = df_processado[df_processado['Cliente'].str.contains(cliente_search, case=False, na=False)]
                            if df_filtrado_cliente.empty:
                                st.warning(f"Nenhum cliente encontrado para '{cliente_search}'. Exibindo visão geral.")
                            else:
                                df_processado = df_filtrado_cliente
                                st.info(f"Exibindo visão filtrada para cliente(s) contendo '{cliente_search}'.")
                else:
                    st.error("A coluna 'Serial' não foi encontrada na base de ativos. O cruzamento não foi realizado.")
            except Exception as e:
                st.error(f"Ocorreu um erro ao ler ou processar o arquivo de ativos: {e}")

    st.subheader("Análise de Posições")

    if 'Modelo de HW' not in df_processado.columns or 'Serial' not in df_processado.columns:
        st.error("As colunas 'Modelo de HW' e/ou 'Serial' não foram encontradas.")
        return

    st.markdown("---")
    modelos_disponiveis = sorted(df_processado['Modelo de HW'].dropna().unique())
    modelos_selecionados = st.multiselect(
        "Selecione o(s) modelo(s) do equipamento:",
        options=modelos_disponiveis
    )

    if modelos_selecionados:
        df_filtrado = df_processado[df_processado['Modelo de HW'].isin(modelos_selecionados)].copy()
    else:
        df_filtrado = df_processado.copy()

    # A lógica de geocodificação foi movida para a função 'process_location_data'

    if 'dias_sem_posicao' in df_filtrado.columns:
        # Definição de modelos GSM
        gsm_models = [
            'MT2000', 'RI0352 - FROTAS 4G', 'MXT-130', 'GS10G', 'MXT-130-P5', 'ST350LC', 'ST380', 
            'RT345BT', 'RI0350 - FROTAS 2G', 'MXT-162', 'E3 Plus', 'ST300H', 'ST340LC', 'SA200', 
            'FMB130', 'GV300CAN'
        ]

        #
        # PASSO 1: Criar o DataFrame de status e aplicar as regras de manutenção
        #
        df_status = df_filtrado.copy()

        # Regra 1: Manutenção por dias sem posição
        manutencao_por_dias = (df_status['dias_sem_posicao'] >= 15) | (df_status['dias_sem_posicao'].isnull())

        # Regra 2: Manutenção para modelos GSM com tecnologia inválida, com exceções
        manutencao_por_tech = pd.Series(False, index=df_status.index)
        if 'tecnologia_posicao' in df_status.columns and 'Modelo de HW' in df_status.columns:
            is_gsm_model = df_status['Modelo de HW'].isin(gsm_models)
            
            # Tecnologias inválidas para GSM são todas que não são 'GSM' (e não são nulas)
            is_invalid_tech = ~df_status['tecnologia_posicao'].isin(['GSM']) & df_status['tecnologia_posicao'].notna()

            # Exceção para modelos como Iotracking e Contingência que podem usar outras tecnologias
            is_exception_model = df_status['Modelo de HW'].str.contains('Iotracking|Contingencia', case=False, na=False)

            # A regra de manutenção só se aplica se NÃO for um modelo de exceção
            manutencao_por_tech = is_gsm_model & is_invalid_tech & ~is_exception_model
        
        # Combinar as regras para definir o status final
        df_status['Status'] = np.where(manutencao_por_dias | manutencao_por_tech, 'Gerar Manutenção', 'OK')
        
        #
        # PASSO 2: Exibir o levantamento de GSMs com tecnologia incorreta (pedido do usuário)
        #
        if 'tecnologia_posicao' in df_filtrado.columns:
            # Usamos df_filtrado para ter a visão antes da regra de manutenção ser aplicada
            df_gsm_wrong_tech = df_filtrado[
                df_filtrado['Modelo de HW'].isin(gsm_models) &
                df_filtrado['tecnologia_posicao'].isin(['Lorawan', 'P2P', 'P2P/LoraWan'])
            ]
            if not df_gsm_wrong_tech.empty:
                st.warning("Levantamento: Modelos GSM que estão posicionando como Lora ou P2P")
                st.dataframe(
                    df_gsm_wrong_tech[['Serial', 'Modelo de HW', 'tecnologia_posicao', 'dias_sem_posicao', 'data_evento']],
                    use_container_width=True,
                    column_config={"data_evento": st.column_config.DatetimeColumn("Data da Posição", format="DD/MM/YYYY HH:mm:ss")}
                )

        #
        # PASSO 3: Apresentar as métricas e a tabela de status principal
        #
        total_equipamentos = len(df_filtrado)
        # Recalcular métricas com base no novo status
        equipamentos_ok = df_status[df_status['Status'] == 'OK']
        
        posicionando_hoje = len(equipamentos_ok[equipamentos_ok['dias_sem_posicao'] <= 1])
        atencao = len(equipamentos_ok[equipamentos_ok['dias_sem_posicao'].between(2, 14.9, inclusive='both')])
        manutencao = len(df_status[df_status['Status'] == 'Gerar Manutenção'])
        sem_posicao = df_filtrado['dias_sem_posicao'].isnull().sum() # Esta métrica não muda

        col1, col2, col3, col4, col5 = st.columns(5)
        col1.metric("Total Dispositivos", total_equipamentos)
        col2.metric("Posicionando OK (<=1d)", posicionando_hoje)
        col3.metric("Atenção (2-14d)", atencao)
        col4.metric("Manutenção", manutencao) # Métrica agora reflete o status
        col5.metric("Nunca Posicionou", sem_posicao)

        st.markdown("---")
        st.subheader("Status de Todos os Equipamentos")
        df_display = df_status.rename(columns={'data_evento': 'Data da Última Posição'})
        colunas_status = ['Serial', 'Cliente', 'Placa', 'Modelo de HW', 'dias_sem_posicao', 'Data da Última Posição', 'Status', 'tecnologia_posicao']
        colunas_existentes = [col for col in colunas_status if col in df_display.columns]
        st.dataframe(df_display[colunas_existentes], use_container_width=True, column_config={"Data da Última Posição": st.column_config.DatetimeColumn("Data da Última Posição", format="DD/MM/YYYY HH:mm:ss")})

        df_status['lat'] = pd.to_numeric(df_status.get('lat'), errors='coerce')
        df_status['long'] = pd.to_numeric(df_status.get('long'), errors='coerce')
        disperso_mask = (~df_status['lat'].between(-34.0, 6.0) | ~df_status['long'].between(-74.0, -34.0)) & df_status['lat'].notna() & df_status['long'].notna()
        df_status['Observação'] = np.where(disperso_mask, 'Disperso', '')

        # --- NOVA SEÇÃO: EXPORTAR EQUIPAMENTOS FORA DO BRASIL ---
        st.markdown("---")
        st.subheader("Equipamentos Fora do Brasil (Posição Divergente)")
        df_fora_brasil = df_status[disperso_mask].copy()

        if not df_fora_brasil.empty:
            st.warning(f"Encontrados {len(df_fora_brasil)} equipamentos com posição fora do Brasil.")

            # --- Reverse Geocoding ---
            paises = []
            if not df_fora_brasil.empty:
                coordenadas = list(zip(df_fora_brasil['lat'], df_fora_brasil['long']))
                resultados_rg = rg.search(coordenadas)
                for res in resultados_rg:
                    try:
                        country = pycountry.countries.get(alpha_2=res['cc'])
                        paises.append(country.name if country else 'Local Desconhecido')
                    except (IndexError, KeyError):
                        paises.append('Oceano / Sem Dados')
            df_fora_brasil['País'] = paises
            
            # Selecionar colunas relevantes para a exportação
            cols_export_fora_brasil = ['Serial', 'Cliente', 'Placa', 'Modelo de HW', 'lat', 'long', 'País', 'dias_sem_posicao', 'Data da Última Posição']
            cols_existentes_fora_brasil = [c for c in cols_export_fora_brasil if c in df_fora_brasil.columns]
            df_export_fora_brasil_display = df_fora_brasil[cols_existentes_fora_brasil]
            
            st.dataframe(df_export_fora_brasil_display, use_container_width=True, column_config={"Data da Última Posição": st.column_config.DatetimeColumn("Data da Última Posição", format="DD/MM/YYYY HH:mm:ss")})

            # --- Mapa de Posições Divergentes ---
            st.subheader("Mapa de Posições Divergentes")
            df_mapa_fora_brasil = df_fora_brasil[['Serial', 'lat', 'long']].copy()
            df_mapa_fora_brasil.rename(columns={'Serial': 'name'}, inplace=True)

            df_mapa_fora_brasil.dropna(subset=['lat', 'long'], inplace=True)
            df_mapa_fora_brasil = df_mapa_fora_brasil[
                (df_mapa_fora_brasil['lat'].between(-90, 90)) &
                (df_mapa_fora_brasil['long'].between(-180, 180)) &
                ((df_mapa_fora_brasil['lat'] != 0) | (df_mapa_fora_brasil['long'] != 0))
            ]

            if not df_mapa_fora_brasil.empty:
                size = st.slider("Raio dos Pontos (em metros)", min_value=1000, max_value=50000, value=5000, step=1000, key="slider_mapa_divergente")
                
                layer = pdk.Layer(
                    "ScatterplotLayer",
                    data=df_mapa_fora_brasil,
                    get_position=['long', 'lat'],
                    get_fill_color=[255, 0, 0, 140],
                    get_radius=size,
                    pickable=True,
                    radius_min_pixels=5,
                )

                view_state = pdk.ViewState(
                    latitude=df_mapa_fora_brasil['lat'].mean(),
                    longitude=df_mapa_fora_brasil['long'].mean(),
                    zoom=2,
                    pitch=0,
                )

                tooltip = {
                    "html": "<b>Serial:</b> {name}",
                    "style": {
                        "backgroundColor": "steelblue",
                        "color": "white",
                    }
                }

                r = pdk.Deck(
                    layers=[layer],
                    initial_view_state=view_state,
                    map_style=pdk.map_styles.DARK,
                    tooltip=tooltip,
                )
                
                st.pydeck_chart(r)
            else:
                st.info("Nenhuma coordenada válida para exibir no mapa de posições divergentes.")


            # Preparar para exportação
            df_export_final = df_fora_brasil[cols_existentes_fora_brasil].copy()
            df_export_final['Serial'] = '="' + df_export_final['Serial'].astype(str) + '"'
            excel_data_fora_brasil = to_excel(df_export_final)
            
            st.download_button(
                label="📥 Baixar Lista de Equipamentos Fora do Brasil (.xlsx)",
                data=excel_data_fora_brasil,
                file_name="equipamentos_fora_do_brasil.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="btn_export_fora_brasil"
            )
        else:
            st.info("Nenhum equipamento com posição registrada fora do Brasil foi encontrado.")
        st.markdown("---")

        df_para_exportar_raw = df_status[df_status['Status'] == 'Gerar Manutenção']
        if not df_para_exportar_raw.empty:
            cols_export = ['Serial', 'Cliente', 'Placa', 'Chassi', 'ObjetoRastreavelStatus', 'Endereco', 'Cidade', 'Uf Proprietario', 'Modelo de HW', 'dias_sem_posicao', 'data_evento', 'Observação', 'tecnologia_posicao']
            cols_export_existentes = [c for c in cols_export if c in df_para_exportar_raw.columns]
            df_export = df_para_exportar_raw[cols_export_existentes].copy()
            df_export.rename(columns={'data_evento': 'Data da Última Posição'}, inplace=True)
            df_export['Serial'] = '="' + df_export['Serial'].astype(str) + '"'
            excel_data = to_excel(df_export)
            st.download_button(label="📥 Baixar Seriais para Manutenção (.xlsx)", data=excel_data, file_name="seriais_manutencao.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

        st.markdown("---")
        st.subheader("Status da Frota por Tempo")
        bins = [-1, 1, 10, 14, 20, 30, np.inf]
        labels = ['Hoje (0-1 dias)', '2-10 dias', '11 a 14 dias', '15 a 20 dias', '21-30 dias', '30+ dias']
        df_filtrado['faixa_dias'] = pd.cut(df_filtrado['dias_sem_posicao'], bins=bins, labels=labels, right=True)
        contagem_faixa = df_filtrado['faixa_dias'].value_counts().sort_index()
        sem_posicao_count = df_filtrado['dias_sem_posicao'].isnull().sum()
        all_labels = labels + ['Sem Posição']
        cols = st.columns(len(all_labels))
        for i, label in enumerate(labels):
            with cols[i]:
                st.metric(label, contagem_faixa.get(label, 0))
        with cols[len(labels)]:
            st.metric("Sem Posição", sem_posicao_count)

    st.markdown("---")
    st.subheader("Quantidade por Modelo de Hardware")
    if 'Modelo de HW' in df_filtrado.columns:
        contagem_modelo = df_filtrado['Modelo de HW'].value_counts()
        num_cols_modelos = 4
        cols_modelos = st.columns(num_cols_modelos)
        col_idx = 0
        for modelo, quantidade in contagem_modelo.items():
            if pd.notna(modelo):
                with cols_modelos[col_idx]:
                    st.metric(str(modelo), quantidade)
                col_idx = (col_idx + 1) % num_cols_modelos
    else:
        st.warning("Coluna 'Modelo de HW' não encontrada.")

    st.markdown("---")
    st.subheader("Detalhes das Últimas Posições")
    colunas_para_mostrar = ['Serial', 'Cliente', 'Placa', 'Status Ativo', 'lat', 'long', 'velocidade', 'tensao bateria', 'cidade', 'odometro', 'odometro_can', 'dias_sem_posicao']
    colunas_existentes = [col for col in colunas_para_mostrar if col in df_filtrado.columns]
    st.dataframe(df_filtrado[colunas_existentes], use_container_width=True)

    cps_cross_df = None
    cps_loaded = "df_cps" in st.session_state and st.session_state.df_cps is not None
    if cps_loaded:
        cps_cross_df = cruzar_odometros_posicao_cps(df_filtrado, st.session_state.get("df_cps"))

    st.markdown("---")
    st.subheader("Cruzamento de Odômetros (Posição x CPS)")
    if not cps_loaded:
        st.info("Carregue um Relatório CPS para habilitar o cruzamento de odômetros.")
    elif cps_cross_df is None:
        st.warning("O relatório CPS precisa conter a coluna 'Evento / Ignição' para extrair os odômetros.")
    elif cps_cross_df.empty:
        st.info("Nenhum equipamento com odômetro cruzado entre Posição e CPS foi encontrado.")
    else:
        st.metric("Combinações encontradas", len(cps_cross_df))
        st.dataframe(cps_cross_df, use_container_width=True)
        st.download_button(
            label="📥 Exportar cruzamento",
            data=convert_df_to_csv(cps_cross_df),
            file_name="cruzamento_odometros_posicao_cps.csv",
            mime="text/csv",
            use_container_width=True
        )

    st.markdown("---")
    st.subheader("Mapa de Posições (Equipamentos posicionando há menos de 15 dias)")
    df_mapa = df_filtrado[df_filtrado['dias_sem_posicao'] < 15].copy()
    df_mapa.dropna(subset=['lat', 'long'], inplace=True)
    if not df_mapa.empty:
        df_mapa = df_mapa[((df_mapa['lat'] != 0) | (df_mapa['long'] != 0)) & (df_mapa['lat'].between(-90, 90)) & (df_mapa['long'].between(-180, 180))]
    if not df_mapa.empty:
        st.info(f"Exibindo {len(df_mapa)} equipamentos com última posição registrada há menos de 15 dias.")
        st.map(df_mapa, latitude='lat', longitude='long')
    else:
        st.warning("Nenhum equipamento posicionando há menos de 15 dias com coordenadas válidas para exibir no mapa.")

    st.markdown("---")
    st.subheader("Mapa de Última Posição (Equipamentos sem posição há mais de 15 dias)")
    df_mapa_antigos = df_filtrado[(df_filtrado['dias_sem_posicao'] >= 15) | (df_filtrado['dias_sem_posicao'].isnull())].copy()
    df_mapa_antigos.dropna(subset=['lat', 'long'], inplace=True)
    if not df_mapa_antigos.empty:
        df_mapa_antigos = df_mapa_antigos[((df_mapa_antigos['lat'] != 0) | (df_mapa_antigos['long'] != 0)) & (df_mapa_antigos['lat'].between(-90, 90)) & (df_mapa_antigos['long'].between(-180, 180))]
    if not df_mapa_antigos.empty:
        st.info(f"Exibindo {len(df_mapa_antigos)} equipamentos com última posição registrada há 15 dias ou mais.")
        st.map(df_mapa_antigos, latitude='lat', longitude='long')
    else:
        st.info("Nenhum equipamento sem posição há mais de 15 dias com coordenadas válidas encontrado.")

    st.markdown("---")
    st.subheader("Modelos por Tecnologia de Posição")
    if 'tecnologia_posicao' in df_filtrado.columns and 'dias_sem_posicao' in df_filtrado.columns:
        df_posicionando = df_filtrado.dropna(subset=['tecnologia_posicao'])
        if not df_posicionando.empty:
            df_posicionando['tecnologia_posicao'] = df_posicionando['tecnologia_posicao'].replace({'P2P': 'P2P/LoraWan'})
            contagem_total = df_posicionando.groupby(['tecnologia_posicao', 'Modelo de HW']).size().reset_index(name='Total')
            df_menor_15 = df_posicionando[df_posicionando['dias_sem_posicao'] < 15]
            contagem_menor_15 = df_menor_15.groupby(['tecnologia_posicao', 'Modelo de HW']).size().reset_index(name='Qtd. < 15 dias')
            df_maior_15 = df_posicionando[(df_posicionando['dias_sem_posicao'] >= 15) | (df_posicionando['dias_sem_posicao'].isnull())]
            contagem_maior_15 = df_maior_15.groupby(['tecnologia_posicao', 'Modelo de HW']).size().reset_index(name='Qtd. >= 15 dias')
            df_final = pd.merge(contagem_total, contagem_menor_15, on=['tecnologia_posicao', 'Modelo de HW'], how='left')
            df_final = pd.merge(df_final, contagem_maior_15, on=['tecnologia_posicao', 'Modelo de HW'], how='left')
            df_final[['Qtd. < 15 dias', 'Qtd. >= 15 dias']] = df_final[['Qtd. < 15 dias', 'Qtd. >= 15 dias']].fillna(0).astype(int)
            tecnologias = df_final['tecnologia_posicao'].unique()
            if len(tecnologias) > 0:
                cols = st.columns(len(tecnologias))
                for i, tech in enumerate(tecnologias):
                    with cols[i]:
                        st.info(f"**{tech}**")
                        df_tech = df_final[df_final['tecnologia_posicao'] == tech].copy()
                        df_tech.set_index('Modelo de HW', inplace=True)
                        df_tech.drop(columns='tecnologia_posicao', inplace=True)
                        df_tech.sort_values(by='Total', ascending=False, inplace=True)
                        st.dataframe(df_tech, use_container_width=True)
            else:
                st.info("Não foram encontradas tecnologias de posição para agrupar os modelos.")
        else:
            st.info("Nenhuma informação de tecnologia de posição encontrada para os equipamentos.")
    else:
        st.info("Colunas essenciais ('tecnologia_posicao', 'dias_sem_posicao') não foram encontradas no relatório processado.")

    st.markdown("---")
    st.subheader("Exportar Lista Completa de Equipamentos")

    # Classificação Capital vs Interior
    if 'cidade_final' in df_filtrado.columns:
        capitais_br = [
            'Rio Branco', 'Maceió', 'Macapá', 'Manaus', 'Salvador', 'Fortaleza', 'Brasília', 'Vitória',
            'Goiânia', 'São Luís', 'Cuiabá', 'Campo Grande', 'Belo Horizonte', 'Belém', 'João Pessoa',
            'Curitiba', 'Recife', 'Teresina', 'Rio de Janeiro', 'Natal', 'Porto Alegre', 'Porto Velho',
            'Boa Vista', 'Florianópolis', 'São Paulo', 'Aracaju', 'Palmas',
            # Variantes sem acento
            'Maceio', 'Macapa', 'Brasilia', 'Vitoria', 'Goiania', 'Sao Luis', 'Cuiaba', 'Belem', 
            'Joao Pessoa', 'Sao Paulo', 'Florianopolis'
        ]
        capitais_set = set(c.lower() for c in capitais_br)
        
        df_filtrado['tipo_localidade'] = df_filtrado['cidade_final'].apply(
            lambda x: 'Capital' if isinstance(x, str) and x.lower() in capitais_set else ('Interior' if pd.notna(x) and str(x).strip() != '' else 'Indefinido')
        )

    # Colunas da Posição para exportar
    posicao_cols = {
        'Serial': 'Serial', 
        'data_evento': 'Data da Posição', 
        'tecnologia_posicao': 'Tecnologia Original',
        'fonte_posicao': 'Fonte da Posição',
        'Modelo de HW': 'Modelo HW',
        'cidade_final': 'Cidade (Posição)',
        'estado_final': 'Estado (Posição)',
        'pais_final': 'País (Posição)',
        'tipo_localidade': 'Tipo Localidade',
        'mesma_regiao': 'Na Região da Base'
    }
    # inclui odômetros nas exportações
    if 'odometro' in df_filtrado.columns:
        posicao_cols['odometro'] = 'Odômetro (Posição)'
    if 'odometro_can' in df_filtrado.columns:
        posicao_cols['odometro_can'] = 'Odômetro CAN (Posição)'
    
    # Colunas dos Ativos para exportar
    ativos_cols = {
        'Codigo': 'Codigo Ativo', 
        'ObjetoRastreavelStatus': 'Status Ativo', 
        'Chassi': 'Chassi', 
        'Placa': 'Placa',
        'Cliente': 'Cliente', 
        'CpfCnpjProprietario': 'CPF/CNPJ', 
        'Endereco': 'Endereço (Base)', 
        'Cidade': 'Cidade (Base)', 
        'Uf Proprietario': 'Estado (Base)',
        'Estado': 'Estado Alternativo (Base)' # Fallback se Uf Proprietario não existir
    }

    if 'Serial' in df_filtrado.columns:
        df_full_export = pd.DataFrame(index=df_filtrado.index)

        # Adiciona colunas da posição
        for df_col, export_name in posicao_cols.items():
            if df_col in df_filtrado.columns:
                df_full_export[export_name] = df_filtrado[df_col]
        
        # Adiciona colunas dos ativos se a base foi cruzada
        for df_col, export_name in ativos_cols.items():
            if df_col in df_filtrado.columns:
                df_full_export[export_name] = df_filtrado[df_col]

        # Garante que as colunas de localização de fallback (do ativo) sejam preenchidas se existirem
        if 'Estado (Base)' not in df_full_export.columns and 'Estado Alternativo (Base)' in df_full_export.columns:
            df_full_export['Estado (Base)'] = df_full_export['Estado Alternativo (Base)']
        df_full_export.drop(columns=['Estado Alternativo (Base)'], inplace=True, errors='ignore')

        # Post-processing on columns
        if 'Serial' in df_full_export.columns:
            df_full_export['Serial'] = '="' + df_full_export['Serial'].astype(str) + '"'
        if 'Tecnologia Original' in df_full_export.columns:
            df_full_export['Tecnologia Original'] = df_full_export['Tecnologia Original'].replace({'P2P': 'P2P/LoraWan'})

        # Agrupamento de múltiplos equipamentos
        if 'Placa' in df_full_export.columns:
            # Usar Placa, Chassi ou Codigo como identificador do veículo
            df_full_export['ID_Veiculo'] = df_full_export['Placa']
            if 'Chassi' in df_full_export.columns:
                df_full_export['ID_Veiculo'] = df_full_export['ID_Veiculo'].fillna(df_full_export['Chassi'])
            if 'Codigo Ativo' in df_full_export.columns:
                df_full_export['ID_Veiculo'] = df_full_export['ID_Veiculo'].fillna(df_full_export['Codigo Ativo'])
            
            # Contar equipamentos por veículo
            counts = df_full_export.groupby('ID_Veiculo')['Serial'].transform('count')
            
            # Adicionar coluna para destacar múltiplos equipamentos
            df_full_export['Multiplos Equipamentos'] = np.where(
                (counts > 1) & df_full_export['ID_Veiculo'].notna(),
                'Sim',
                'Não'
            )
            
            # Adicionar coluna de grupo e ordenar
            df_full_export['Grupo'] = df_full_export['ID_Veiculo'].where(counts > 1)
            df_full_export.sort_values(by=['Grupo', 'Serial'], na_position='last', inplace=True)
            df_full_export.drop(columns=['ID_Veiculo'], inplace=True)
        else:
            # Se não houver dados de ativos, não há agrupamento
            df_full_export['Multiplos Equipamentos'] = 'Não'
            df_full_export['Grupo'] = pd.NA

        # Split for download buttons
        if 'Data da Posição' in df_full_export.columns:
            df_com_posicao = df_full_export[df_full_export['Data da Posição'].notna()].copy()
            df_sem_posicao = df_full_export[df_full_export['Data da Posição'].isna()].copy()
        else:
            df_com_posicao = pd.DataFrame(columns=df_full_export.columns)
            df_sem_posicao = df_full_export.copy()

        col1, col2, col3 = st.columns(3)
        with col1:
            if not df_com_posicao.empty:
                st.download_button(label="📥 Baixar COM Posição", data=to_excel(df_com_posicao, highlight_col='Multiplos Equipamentos'), file_name="equipamentos_com_posicao.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            else:
                st.info("Nenhum equipamento com posição para exportar.")
        with col2:
            if not df_sem_posicao.empty:
                df_sem_posicao_export = df_sem_posicao.drop(columns=['Data da Posição', 'Tecnologia'], errors='ignore')
                st.download_button(label="📥 Baixar SEM Posição", data=to_excel(df_sem_posicao_export, highlight_col='Multiplos Equipamentos'), file_name="equipamentos_sem_posicao.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            else:
                st.info("Nenhum equipamento sem posição para exportar.")
        with col3:
            if not df_full_export.empty:
                st.download_button(label="📥 Baixar TODOS Equipamentos", data=to_excel(df_full_export, highlight_col='Multiplos Equipamentos'), file_name="equipamentos_todos.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            else:
                st.info("Nenhum equipamento para exportar.")
