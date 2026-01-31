# app.py
import streamlit as st
import pandas as pd
import numpy as np
from haversine import haversine, Unit
import io
from openpyxl.styles import PatternFill
import streamlit_authenticator as stauth
import bcrypt
import yaml
from yaml.loader import SafeLoader

# Importar todos os módulos da aplicação
from modules.config import configurar_app, carregar_chave_api
from modules.posicao import analisar_ultima_posicao
from modules.cps import analisar_cps
from modules.session import inicializar_sessao
from modules.data_loader import (
    uploader_agendamentos, uploader_mapeamento, uploader_pagamento, uploader_backlog, uploader_ultimaposicao,
    uploader_devolucao, uploader_ativos, uploader_cps, uploader_ordens_pendentes, limpar_tudo
)
from modules.dashboard import exibir_dashboard
from modules.custos import analisar_custos
from modules.devolucao import ferramenta_devolucao
from modules.mapeamento import ferramenta_mapeamento
from modules.otimizador import otimizador
from modules.ativos import ferramenta_ativos
from modules.chat import chat_interface
from modules.agendadas import exibir_ordens_agendadas
from modules.distancia import analisar_distancia_percorrida  # Importa a nova função
from modules.utils import (
    executar_analise_segura as executar_analise_pandas_fn,
    convert_df_to_csv, 
    safe_to_numeric
)

# Importa a nova função de processamento que criamos
from modules.processar_relatorio import processar_dataframe_posicao

# --- FUNÇÃO PARA ESTILIZAÇÃO CSS ---
def inject_custom_css():
    """Injeta CSS customizado para melhorar a aparência da aplicação."""
    st.markdown("""
        <style>
            /* Melhora a aparência geral dos botões */
            .stButton > button {
                border-radius: 8px;
                border: 1px solid transparent;
                transition: all 0.3s ease-in-out;
                box-shadow: 0 1px 3px rgba(0,0,0,0.12), 0 1px 2px rgba(0,0,0,0.24);
            }
            .stButton > button:hover {
                box-shadow: 0 3px 6px rgba(0,0,0,0.16), 0 3px 6px rgba(0,0,0,0.23);
                transform: translateY(-1px);
            }

            /* Estilo para abas */
            .stTabs [data-baseweb="tab-list"] {
                gap: 2px;
            }
            .stTabs [data-baseweb="tab"] {
                height: 40px;
                white-space: pre-wrap;
                background-color: #262730; /* Cor de fundo da aba inativa */
                border-radius: 4px 4px 0px 0px;
                gap: 1px;
                padding-top: 10px;
                padding-bottom: 10px;
            }
            .stTabs [aria-selected="true"] {
                background-color: #3A3A3A; /* Cor de fundo da aba ativa */
            }
        </style>
    """, unsafe_allow_html=True)

# --- FUNÇÕES DO BACKLOG (MOVENDO PARA DENTRO DO APP.PY) ---

def to_styled_excel(df: pd.DataFrame) -> bytes:
    """
    Converte um DataFrame para um arquivo Excel (.xlsx) em memória,
    com formatação e estilo aplicados.
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Arredonda a coluna de distância para 1 casa decimal
        df_export = df.copy()
        if 'DISTANCIA_KM' in df_export.columns:
            df_export['DISTANCIA_KM'] = df_export['DISTANCIA_KM'].round(1)
        
        df_export.to_excel(writer, index=False, sheet_name='Resultado')
        
        # Pega a worksheet para aplicar o estilo
        workbook = writer.book
        worksheet = writer.sheets['Resultado']
        
        # Define o preenchimento cinza claro
        light_gray_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # Colunas para colorir
        cols_to_style = ['RANKING', 'REPRESENTANTE', 'CIDADE_RT', 'DISTANCIA_KM', 'OS']
        
        # Itera pelas colunas do DataFrame para encontrar o índice correto
        for col_name in cols_to_style:
            if col_name in df_export.columns:
                # +1 porque as colunas no openpyxl são 1-based
                col_idx = df_export.columns.get_loc(col_name) + 1
                for cell in worksheet.iter_cols(min_col=col_idx, max_col=col_idx):
                    for c in cell:
                        c.fill = light_gray_fill

    return output.getvalue()


def processar_backlog_df(df_backlog, df_mapeamento, num_rts_proximos=2):
    """
    Processa o DataFrame de backlog para encontrar os N RTs mais próximos para cada O.S.
    """
    # --- FILTRO DE REPRESENTANTES ESPECIAIS ---
    # Identifica a coluna de representante no mapeamento para o filtro
    rep_col_filtro = next((c for c in df_mapeamento.columns if 'nm_representante' in c.lower()), None)
    if rep_col_filtro:
        palavras_excluir = ['STELLANTIS', 'CEABS']
        df_mapeamento = df_mapeamento[~df_mapeamento[rep_col_filtro].str.contains('|'.join(palavras_excluir), case=False, na=False)].copy()

    # Lógica aprimorada para encontrar colunas
    os_col_b = next((c for c in df_backlog.columns if c.lower() in ['os', 'numero os', 'ordem de servico', 'ordem', 'ordemservicoid']), None)
    
    if not os_col_b:
        st.error("Coluna de 'OS' não encontrada no arquivo de backlog. Verifique se o nome da coluna é 'OS', 'Numero OS' ou similar.")
        return pd.DataFrame()
    
    # Garante que a coluna de OS seja sempre chamada 'OS' para consistência
    df_backlog.rename(columns={os_col_b: 'OS'}, inplace=True)
    os_col_b = 'OS' # Atualiza a variável para o novo nome padrão

    lat_col_b = next((c for c in df_backlog.columns if 'latitude' in c.lower()), None)
    lon_col_b = next((c for c in df_backlog.columns if 'longitude' in c.lower()), None)
    coord_col_b = next((c for c in df_backlog.columns if any(k in c.lower() for k in ['coord', 'lat/long', 'latlong'])), None)

    # Se não encontrar lat/lon separadas, tenta a coluna unificada
    if not lat_col_b and not lon_col_b and coord_col_b:
        try:
            coords = df_backlog[coord_col_b].astype(str).str.strip().str.replace('"', '').str.split(',', expand=True)
            df_backlog['lat_backlog'] = safe_to_numeric(coords[0])
            df_backlog['lon_backlog'] = safe_to_numeric(coords[1])
        except Exception as e:
            st.error(f"Falha ao processar a coluna de coordenadas '{coord_col_b}': {e}")
            return pd.DataFrame()
    elif lat_col_b and lon_col_b:
        df_backlog['lat_backlog'] = safe_to_numeric(df_backlog[lat_col_b])
        df_backlog['lon_backlog'] = safe_to_numeric(df_backlog[lon_col_b])
    else:
        # --- NOVA LÓGICA DE FALLBACK: Buscar coordenadas pela cidade no Mapeamento ---
        st.info("Coordenadas não encontradas no backlog. Tentando buscar pela cidade no arquivo de Mapeamento...")
        
        cidade_col_b = next((c for c in df_backlog.columns if 'cidade' in c.lower() and 'rt' not in c.lower()), None)
        if not cidade_col_b:
            st.error("Não foi possível encontrar colunas de 'Latitude'/'Longitude' nem uma coluna de 'Cidade' no arquivo de backlog.")
            return pd.DataFrame()

        # Preparar o mapeamento para consulta (Cidade -> Coordenadas)
        map_city_col = next((c for c in df_mapeamento.columns if 'nm_cidade_atendimento' in c.lower()), None)
        map_lat_col = next((c for c in df_mapeamento.columns if 'cd_latitude_atendimento' in c.lower()), None)
        map_lon_col = next((c for c in df_mapeamento.columns if 'cd_longitude_atendimento' in c.lower()), None)

        if not all([map_city_col, map_lat_col, map_lon_col]):
            st.error("O arquivo de Mapeamento não contém as colunas necessárias ('nm_cidade_atendimento', 'cd_latitude_atendimento', 'cd_longitude_atendimento') para a busca por cidade.")
            return pd.DataFrame()

        # Cria um dicionário de busca removendo duplicatas
        df_mapeamento_coords = df_mapeamento[[map_city_col, map_lat_col, map_lon_col]].dropna().drop_duplicates(subset=[map_city_col])
        df_mapeamento_coords[map_city_col] = df_mapeamento_coords[map_city_col].str.upper().str.strip()
        coord_lookup = df_mapeamento_coords.set_index(map_city_col).to_dict('index')

        # Aplica o lookup para obter as coordenadas
        df_backlog['coords'] = df_backlog[cidade_col_b].str.upper().str.strip().map(coord_lookup)
        df_backlog['lat_backlog'] = df_backlog['coords'].apply(lambda x: x[map_lat_col] if isinstance(x, dict) else None)
        df_backlog['lon_backlog'] = df_backlog['coords'].apply(lambda x: x[map_lon_col] if isinstance(x, dict) else None)
        df_backlog.drop(columns=['coords'], inplace=True)

    df_backlog = df_backlog.dropna(subset=['lat_backlog', 'lon_backlog'])
    if df_backlog.empty:
        st.warning("Nenhuma ordem de serviço com coordenadas válidas foi encontrada após a limpeza e cruzamento com o Mapeamento.")
        return pd.DataFrame()

    df_backlog = df_backlog.dropna(subset=['lat_backlog', 'lon_backlog'])
    if df_backlog.empty:
        st.warning("Nenhuma ordem de serviço com coordenadas válidas foi encontrada após a limpeza dos dados.")
        return pd.DataFrame()

    lat_col_m = next((c for c in df_mapeamento.columns if 'cd_latitude_representante' in c.lower()), None)
    lon_col_m = next((c for c in df_mapeamento.columns if 'cd_longitude_representante' in c.lower()), None)
    rep_col_m = next((c for c in df_mapeamento.columns if 'nm_representante' in c.lower()), None)
    cidade_rt_col_m = next((c for c in df_mapeamento.columns if 'nm_cidade_representante' in c.lower()), None)

    if not all([lat_col_m, lon_col_m, rep_col_m, cidade_rt_col_m]):
        st.error("Arquivo de mapeamento precisa conter colunas de 'Representante', 'Cidade RT', 'Latitude' e 'Longitude'.")
        return pd.DataFrame()

    df_mapeamento['lat_rt'] = safe_to_numeric(df_mapeamento[lat_col_m])
    df_mapeamento['lon_rt'] = safe_to_numeric(df_mapeamento[lon_col_m])
    df_mapeamento = df_mapeamento.dropna(subset=['lat_rt', 'lon_rt'])
    
    rts_unicos = df_mapeamento[[rep_col_m, cidade_rt_col_m, 'lat_rt', 'lon_rt']].drop_duplicates(subset=[rep_col_m])

    resultados = []
    for _, os_row in df_backlog.iterrows():
        os_coord = (os_row['lat_backlog'], os_row['lon_backlog'])
        distancias = []
        for _, rt_row in rts_unicos.iterrows():
            rt_coord = (rt_row['lat_rt'], rt_row['lon_rt'])
            dist = haversine(os_coord, rt_coord, unit=Unit.KILOMETERS)
            distancias.append({'OS': os_row[os_col_b], 'REPRESENTANTE': rt_row[rep_col_m], 'CIDADE_RT': rt_row[cidade_rt_col_m], 'DISTANCIA_KM': dist})
        
        distancias_df = pd.DataFrame(distancias)
        mais_proximos = distancias_df.nsmallest(num_rts_proximos, 'DISTANCIA_KM')
        mais_proximos['RANKING'] = range(1, len(mais_proximos) + 1)
        
        os_info = os_row.drop(labels=['lat_backlog', 'lon_backlog'])
        for _, proximo_row in mais_proximos.iterrows():
            resultado_final = proximo_row.to_dict()
            resultado_final.update(os_info.to_dict())
            resultados.append(resultado_final)

    if not resultados: return pd.DataFrame()
    df_final = pd.DataFrame(resultados)
    
    cols_principais = ['RANKING', 'REPRESENTANTE', 'CIDADE_RT', 'DISTANCIA_KM']
    cols_backlog = [c for c in df_backlog.columns if c not in ['lat_backlog', 'lon_backlog']]
    ordem_final = [c for c in cols_principais + cols_backlog if c in df_final.columns]
    
    # Garante que a coluna 'OS' exista, renomeando a coluna original se necessário.
    os_col_original = next((c for c in df_final.columns if c.lower() in ['os', 'numero os', 'ordem de servico', 'ordem', 'ordemservicoid']), None)
    if os_col_original and os_col_original != 'OS':
        df_final.rename(columns={os_col_original: 'OS'}, inplace=True)
    return df_final[ordem_final]

def render_backlog_processor():
    df_mapeamento = st.session_state.get("df_mapeamento")
    df_backlog = st.session_state.get("df_backlog")

    st.markdown("---"); st.subheader("2. Filtros e Processamento")
    col1, col2, col3 = st.columns([2, 2, 1])
    uf_col = next((c for c in df_backlog.columns if 'uf' in c.lower()), None)
    cidade_col = next((c for c in df_backlog.columns if 'cidade' in c.lower() and 'rt' not in c.lower()), None)
    df_filtrado = df_backlog.copy()

    if uf_col:
        ufs = sorted(df_filtrado[uf_col].dropna().unique())
        if uf_selecionada := col1.multiselect("Filtrar por UF:", options=ufs):
            df_filtrado = df_filtrado[df_filtrado[uf_col].isin(uf_selecionada)]
    if cidade_col:
        cidades = sorted(df_filtrado[cidade_col].dropna().unique())
        if cidade_selecionada := col2.multiselect("Filtrar por Cidade:", options=cidades):
            df_filtrado = df_filtrado[df_filtrado[cidade_col].isin(cidade_selecionada)]

    num_rts = col3.number_input("Nº de RTs próximos:", min_value=1, max_value=10, value=2, step=1)

    if st.button("Processar Backlog", use_container_width=True, type="primary"):
        if df_filtrado.empty: st.warning("Nenhum dado no backlog corresponde aos filtros selecionados.")
        else:
            with st.spinner(f"Processando {len(df_filtrado)} ordens e buscando os {num_rts} RTs mais próximos..."):
                df_resultado = processar_backlog_df(df_filtrado, df_mapeamento, num_rts)
                st.session_state.df_backlog_resultado = df_resultado if not df_resultado.empty else None

    if (df_resultado := st.session_state.get("df_backlog_resultado")) is not None:
        st.markdown("---"); st.subheader("3. Resultados")
        st.success(f"Análise concluída! Encontrados RTs para {df_resultado['OS'].nunique()} ordens de serviço.")
        df_display = df_resultado.copy()
        if 'DISTANCIA_KM' in df_display.columns:
            df_display['DISTANCIA_KM'] = df_display['DISTANCIA_KM'].map('{:,.1f} km'.format)
        st.dataframe(df_display, use_container_width=True)
        
        # Novo botão de download para Excel
        st.download_button(
            label="📥 Exportar Resultado (.xlsx)",
            data=to_styled_excel(df_resultado),
            file_name="backlog_rts_proximos.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# --- FIM DAS FUNÇÕES DO BACKLOG ---

# --- LÓGICA DE AUTENTICAÇÃO E REGISTRO ---
with open('config.yaml', encoding='utf-8') as file:
    config = yaml.load(file, Loader=SafeLoader)

authenticator = stauth.Authenticate(
    config['credentials'],
    config['cookie']['name'],
    config['cookie']['key'],
    config['cookie']['expiry_days']
)

auth_status = st.session_state.get("authentication_status")
if auth_status in (None, False):
    _, col_login, col_logo, _ = st.columns([0.4, 1, 1.4, 0.4])
    with col_login:
        authenticator.login()
    with col_logo:
        st.image("assets/ChatGPT Image 26 de jan. de 2026, 10_08_23.png", width=380)
else:
    authenticator.login()

if st.session_state["authentication_status"]:
    # Recarrega o config para ter a informação mais recente
    with open('config.yaml', encoding='utf-8') as file:
        config = yaml.load(file, Loader=SafeLoader)
    
    username = st.session_state.get("username")
    user_details = config['credentials']['usernames'].get(username, {})

    # VERIFICAÇÃO ADICIONAL DE STATUS
    if user_details.get('status') == 'approved':
        # --- APLICAÇÃO PRINCIPAL (SÓ RODA SE AUTENTICADO E APROVADO) ---
        # Configura app e carrega modelo (Início da execução principal)
        configurar_app()
        model = carregar_chave_api() # A API Key é carregada

        # Injeta o CSS customizado para um visual mais profissional
        inject_custom_css()

        # Inicializa sessão com o modelo
        inicializar_sessao(model)

        # Registra a função de análise em session_state para o chat usar
        st.session_state.executar_analise_pandas_fn = executar_analise_pandas_fn

        # --- BARRA LATERAL (SIDEBAR) ---
        with st.sidebar:
            st.sidebar.write(f'Bem-vindo(a), *{st.session_state["name"]}*')
            authenticator.logout()

            # O "Desenvolvido por..." é carregado pela função carregar_chave_api()
            data_keys = [
                "df_agendamentos", "df_ativos", "df_mapeamento", 
                "df_pagamento", "df_devolucao", "df_ultimaposicao", "df_cps", "df_ordens_pendentes"
            ]
            is_loaded = any(st.session_state.get(key) is not None for key in data_keys)
            
            if is_loaded:
                data_status = "Status: Pronto"
                color = "#00C853" # Verde
            else:
                data_status = "Status: Aguardando"
                color = "#0091EA" # Azul
            
            st.sidebar.markdown(
                f"""
                <div style="padding: 12px; border-radius: 8px; background-color: #262730; text-align: left; margin-bottom: 20px; border-left: 5px solid {color};">
                    <span style="font-size: 16px; font-weight: bold; color: #FAFAFA;">
                        {data_status.split(':')[0]}:
                    </span>
                    <span style="font-size: 16px; color: {color}; font-weight: bold; float: right;">
                        {data_status.split(':')[1].strip()}
                    </span>
                </div>
                """, 
                unsafe_allow_html=True
            )
            
            if st.sidebar.button("🧹 Limpar Tudo", use_container_width=True):
                limpar_tudo()
        # --- FIM DA BARRA LATERAL ---

        # --- LAYOUT CO-PILOTO ---
        with st.container():
            col_main, col_chat = st.columns([2, 1])

            with col_main:
                # --- LÓGICA DE RENDERIZAÇÃO DINÂMICA DE ABAS ---
                
                # 1. Define todas as abas possíveis na aplicação (incluindo a nova aba de Viagens)
                todas_as_abas_disponiveis = ["📊 Dashboard", "\U0001F5D3\uFE0F Agendadas", "📋 Backlog", "📍 Posição", "✈️ Viagens", "🛰️ CPS", "🚚 Ativos", "⚙️ Otimizador", "💰 Custos", "↩️ Devolução", "🗺️ Mapeamento", "🤖 Co-piloto"]
                
                # 2. Inicializa a lista de abas a serem renderizadas
                abas_para_renderizar = []

                # 3. Lógica para admin ou usuário comum
                if st.session_state.get("username") == 'admin':
                    # Admin vê todas as abas disponíveis
                    abas_para_renderizar = todas_as_abas_disponiveis.copy()
                    abas_para_renderizar.insert(0, "👤 Meu Perfil") # Perfil para o admin
                    abas_para_renderizar.append("👑 Admin")
                else:
                    # Usuário comum vê as abas permitidas em seu perfil
                    user_abas_permitidas = user_details.get('tabs', [])
                    abas_para_renderizar = [aba for aba in todas_as_abas_disponiveis if aba in user_abas_permitidas]
                    abas_para_renderizar.insert(0, "👤 Meu Perfil") # Perfil para todos os logados

                # 4. Renderiza as abas ou uma mensagem de aviso
                if not abas_para_renderizar:
                    st.warning("Você não tem permissão para acessar nenhuma funcionalidade. Por favor, contate um administrador.")
                else:
                    created_tabs = st.tabs(abas_para_renderizar)
                    
                    # 5. Mapeia o nome da aba para o seu conteúdo e renderiza
                    for i, tab_name in enumerate(abas_para_renderizar):
                        with created_tabs[i]:
                            try:
                                # Adiciona um log de acesso
                                log_timestamp = pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')
                                st.session_state.app_log.append(f"[{log_timestamp}] INFO: Usuário '{st.session_state.get('username')}' acessou a aba '{tab_name}'.")

                                if tab_name == "👤 Meu Perfil":
                                    st.subheader(f"Perfil de {st.session_state.get('name')}")
                                    st.write(f"**Nome de Usuário:** `{st.session_state.get('username')}`")
                                    st.write(f"**Email:** `{user_details.get('email')}`")
                                    
                                    st.divider()

                                    with st.expander("🔑 Alterar minha senha"):
                                        try:
                                            if authenticator.reset_password(st.session_state.get("username")):
                                                with open('config.yaml', 'w', encoding='utf-8') as file:
                                                    yaml.dump(config, file, default_flow_style=False)
                                                st.success("Sua senha foi alterada com sucesso!")
                                        except Exception as e:
                                            st.error(e)
                                
                                elif tab_name == "📊 Dashboard":
                                    st.subheader("Carregar Dados de O.S. (Agendamentos)")
                                    uploader_agendamentos()
                                    st.markdown("---")
                                    if "df_agendamentos" in st.session_state and st.session_state.df_agendamentos is not None:
                                        st.success(f"Arquivo de agendamentos carregado com {len(st.session_state.df_agendamentos)} linhas. Clique abaixo para gerar o dashboard.")
                                        if st.button("📊 Gerar Dashboard de Agendamentos", use_container_width=True, type="primary"):
                                            with st.spinner("Analisando dados e gerando gráficos... Por favor, aguarde."):
                                                exibir_dashboard(st.session_state.df_agendamentos)
                                    else:
                                        st.info("👆 Por favor, carregue a 'Pesquisa de O.S.' acima para habilitar o Dashboard.")
                                
                                elif tab_name == "\U0001F5D3\uFE0F Agendadas":
                                    st.subheader("Carregar Ordens Pendentes")
                                    uploader_ordens_pendentes()
                                    st.markdown("---")
                                    exibir_ordens_agendadas(st.session_state.get("df_ordens_pendentes"))

                                elif tab_name == "📋 Backlog":
                                    st.subheader("1. Carregar Arquivo de Backlog")
                                    uploader_backlog(key="main_backlog_uploader")
                                    if st.session_state.get("df_mapeamento") is None:
                                        st.warning("👈 Por favor, carregue o 'Mapeamento de RT' na aba 'Otimizador' para habilitar o processamento do backlog.")
                                    elif st.session_state.get("df_backlog") is None:
                                        st.info("👆 Por favor, carregue o arquivo de 'Backlog' acima para iniciar a análise.")
                                    else:
                                        render_backlog_processor()

                                elif tab_name == "📍 Posição":
                                    st.subheader("Carregar Relatório de Última Posição")
                                    uploader_ultimaposicao()
                                    st.markdown("---")
                                    if "df_ultimaposicao" in st.session_state and st.session_state.df_ultimaposicao is not None:
                                        df_bruto = st.session_state.df_ultimaposicao
                                        with st.spinner("Extraindo dados de geolocalização..."):
                                            df_processado = processar_dataframe_posicao(df_bruto)
                                        analisar_ultima_posicao(df_processado)

                                elif tab_name == "✈️ Viagens":
                                    analisar_distancia_percorrida()

                                elif tab_name == "🛰️ CPS":
                                    st.subheader("Carregar Relatório CPS")
                                    uploader_cps()
                                    st.markdown("---")
                                    if "df_cps" in st.session_state and st.session_state.df_cps is not None:
                                        analisar_cps(st.session_state.df_cps)

                                elif tab_name == "🚚 Ativos":
                                    st.subheader("Carregar Base de Ativos")
                                    uploader_ativos()
                                    st.markdown("---")
                                    if "df_ativos" in st.session_state and st.session_state.df_ativos is not None:
                                        ferramenta_ativos(st.session_state.df_ativos)

                                elif tab_name == "⚙️ Otimizador":
                                    st.subheader("Carregar Mapeamento de RT")
                                    uploader_mapeamento()
                                    st.markdown("---")
                                    if "df_agendamentos" not in st.session_state or st.session_state.df_agendamentos is None:
                                        st.warning("👈 Por favor, carregue a 'Pesquisa de O.S.' na aba 'Dashboard' primeiro.")
                                    elif "df_mapeamento" not in st.session_state or st.session_state.df_mapeamento is None:
                                        st.info("👆 Por favor, carregue o 'Mapeamento de RT' acima para usar o Otimizador.")
                                    else:
                                        otimizador(st.session_state.df_agendamentos, st.session_state.df_mapeamento, st.session_state.get("df_ultimaposicao"), st.session_state.get("df_ativos"))

                                elif tab_name == "💰 Custos":
                                    st.subheader("Carregar Base de Pagamento")
                                    uploader_pagamento()
                                    st.markdown("---")
                                    if "df_pagamento" in st.session_state and st.session_state.df_pagamento is not None:
                                        df_agendamentos = st.session_state.get("df_agendamentos", None)
                                        df_mapeamento = st.session_state.get("df_mapeamento", None)
                                        analisar_custos(st.session_state.df_pagamento, df_agendamentos, df_mapeamento)
                                    else:
                                        st.info("👆 Por favor, carregue a 'Base de Pagamento' acima para analisar os custos.")

                                elif tab_name == "↩️ Devolução":
                                    st.subheader("Carregar Base de Devolução")
                                    uploader_devolucao()
                                    st.markdown("---")
                                    if "df_devolucao" in st.session_state and st.session_state.df_devolucao is not None:
                                        ferramenta_devolucao(st.session_state.df_devolucao)

                                elif tab_name == "🗺️ Mapeamento":
                                    if "df_mapeamento" in st.session_state and st.session_state.df_mapeamento is not None:
                                        ferramenta_mapeamento(st.session_state.df_mapeamento)
                                    else:
                                        st.warning("👈 Por favor, carregue o 'Mapeamento de RT' na aba 'Otimizador' para ver o Mapa.")
                                
                                elif tab_name == "🤖 Co-piloto":
                                    st.header("Informações de Contexto para o Co-piloto")
                                    st.info("O Mercúrio usa o contexto dos dados carregados para responder suas perguntas. Não é necessário copiar e colar.")
                                    for key, name in {"df_ultimaposicao": "Posição", "df_agendamentos": "Agendamentos (Dashboard)", "df_backlog": "Backlog", "df_pagamento": "Custos", "df_devolucao": "Devolução", "df_mapeamento": "Mapeamento", "df_ativos": "Ativos"}.items():
                                        if key in st.session_state and st.session_state[key] is not None:
                                            df = st.session_state[key]
                                            st.subheader(f" Aba: {name}")
                                            st.text(f"  - Total de linhas: {len(df)}")
                                            st.text(f"  - Colunas: {', '.join(df.columns)}")
                                        else:
                                            st.warning(f"Nenhum dado carregado para a aba: {name}")
                                    
                                    # --- NOVO PAINEL DE LOG PARA ADMINS ---
                                    if st.session_state.get("username") == 'admin':
                                        st.markdown("---")
                                        st.subheader("🕵️‍♂️ Log de Atividades e Erros (Admin)")
                                        with st.expander("Clique para ver os logs da sessão atual"):
                                            if not st.session_state.app_log:
                                                st.info("Nenhuma atividade registrada nesta sessão ainda.")
                                            else:
                                                log_text = "\n".join(reversed(st.session_state.app_log))
                                                st.text_area("Logs:", log_text, height=300, key="log_display_admin")


                                elif tab_name == "👑 Admin":
                                    st.subheader("Painel de Administração de Usuários")
                                    with open('config.yaml', encoding='utf-8') as file:
                                        current_config = yaml.load(file, Loader=SafeLoader)
                                    usernames_to_manage = {u: d for u, d in current_config['credentials']['usernames'].items() if u != st.session_state.get("username")}
                                    if not usernames_to_manage:
                                        st.info("Nenhum outro usuário no sistema para gerenciar.")
                                    else:
                                        for username, details in usernames_to_manage.items():
                                            st.markdown("---")
                                            status = "Aprovado" if details.get('status') == 'approved' else "Pendente"
                                            st.markdown(f"#### Usuário: `{username}` (`{status}`)")
                                            with st.form(key=f"form_{username}"):
                                                st.write("**Permissões e Status:**")
                                                
                                                abas_atuais = details.get('tabs', [])
                                                abas_selecionadas = st.multiselect("Selecione as abas:", options=todas_as_abas_disponiveis, default=abas_atuais, key=f"multiselect_{username}")
                                                
                                                status_options = ["approved", "pending", "disabled"]
                                                current_status_index = status_options.index(status) if status in status_options else 0
                                                new_status = st.selectbox("Status da Conta:", options=status_options, index=current_status_index, key=f"status_{username}")

                                                st.divider()
                                                st.write("**Redefinir Senha:**")
                                                new_password = st.text_input("Nova Senha (deixe em branco para não alterar)", type="password", key=f"new_pw_{username}")

                                                submitted = st.form_submit_button("Salvar Todas as Alterações")
                                                
                                                if submitted:
                                                    with open('config.yaml', encoding='utf-8') as file:
                                                        update_config = yaml.load(file, Loader=SafeLoader)

                                                    if username in update_config['credentials']['usernames']:
                                                        # Atualizar abas
                                                        update_config['credentials']['usernames'][username]['tabs'] = abas_selecionadas
                                                        
                                                        # Atualizar status
                                                        update_config['credentials']['usernames'][username]['status'] = new_status
                                                        
                                                        # Atualizar senha, se uma nova foi fornecida
                                                        if new_password:
                                                            hashed_password = bcrypt.hashpw(new_password.encode(), bcrypt.gensalt()).decode()
                                                            update_config['credentials']['usernames'][username]['password'] = hashed_password
                                                        
                                                        with open('config.yaml', 'w', encoding='utf-8') as file:
                                                            yaml.dump(update_config, file, default_flow_style=False)
                                                        
                                                        st.success(f"Informações do usuário '{username}' atualizadas com sucesso!")
                                                        st.rerun()
                                                    else:
                                                        st.error(f"Usuário '{username}' não encontrado no arquivo de configuração ao tentar salvar.")
                                            
                                            # Removida a zona de perigo para exclusão permanente, favorecendo a desativação.


                            except Exception as e:
                                log_timestamp_error = pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')
                                error_message = f"[{log_timestamp_error}] ERROR: Erro na aba '{tab_name}': {str(e)}"
                                st.session_state.app_log.append(error_message)
                                st.error(f"Ocorreu um erro inesperado na aba '{tab_name}'. O erro foi registrado para análise do administrador.")
                                st.exception(e)
            with col_chat:
                chat_interface()
    
    else: # Usuário autenticou, mas não está aprovado
        st.warning('Sua conta foi criada com sucesso e agora está aguardando a aprovação de um administrador.')
        st.info('Você será desconectado. Por favor, aguarde a liberação para acessar o sistema.')
        authenticator.logout()

elif st.session_state["authentication_status"] is False:
    st.markdown("<div style='text-align:center;'><strong>Usuário ou senha incorreta.</strong></div>", unsafe_allow_html=True)
    st.error("")

elif st.session_state["authentication_status"] is None:
    st.markdown("<div style='text-align:center;'><strong>Por favor, insira seu usuário e senha para ter acesso ao Mercúrio.</strong></div>", unsafe_allow_html=True)
    st.warning("")
    _, col_registro, _, _ = st.columns([0.4, 1, 1.4, 0.4])
    # --- FORMULÁRIO DE REGISTRO (DENTRO DE UM EXPANDER) ---
    with col_registro:
        with st.expander("Não tem uma conta? Registre-se aqui"):
            try:
                # Armazena os nomes de usuário existentes antes do registro
                usernames_before = set(config['credentials']['usernames'].keys())

                if authenticator.register_user(fields={
                    'Form name': 'Formulário de Novo Usuário',
                    'Name': 'Nome completo',
                    'Email': 'Email',
                    'Username': 'Nome de usuário',
                    'Password': 'Senha',
                    'Repeat Password': 'Repetir Senha',
                    'Password hint': 'Dica de senha',
                    'Captcha': 'Captcha'
                }):
                    # Encontra o novo nome de usuário
                    usernames_after = set(config['credentials']['usernames'].keys())
                    new_username = (usernames_after - usernames_before).pop()

                    # Define o status como 'approved' e as abas padrão
                    config['credentials']['usernames'][new_username]['status'] = 'approved'
                    config['credentials']['usernames'][new_username]['tabs'] = ["📊 Dashboard", "⚙️ Otimizador"]
                    
                    # Atualiza o config no disco
                    with open('config.yaml', 'w', encoding='utf-8') as file:
                        yaml.dump(config, file, default_flow_style=False)
                    
                    st.success('Usuário registrado e aprovado com sucesso! Você já pode fazer o login.')
                    st.balloons() # Adiciona um pouco de comemoração :)
            except Exception as e:
                st.error(e)
 
